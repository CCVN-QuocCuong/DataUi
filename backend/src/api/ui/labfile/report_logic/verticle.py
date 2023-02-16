from openpyxl import Workbook
from openpyxl.styles import PatternFill
from src.api.ui.labfile.report_logic.report_common import get_date_from_sample_name
from src.api.ui.labfile.report_logic.report_common import (
    get_lab_name_report_by_cocid,
    get_ttcl_labfiledetail_by_cocid_barcode,
)
from src.api.ui.labfile.report_logic.report_common import apply_smart_criteria
from src.api.ui.labfile.report_logic.report_common import ( 
    get_lab_sample_name_by_cocid,
    generate_notes_footer_vertical,
    get_rows_template_excel,
    get_columns_criteria,
    get_datas_criteria,
    get_indicated_asbestos_value, 
    get_notecode_style,
    clean_value,
)
from src.api.ui.labfile.report_logic.report_style import custom_style
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from copy import copy

PAGESIZE = {"A4": {"width": 210, "height": 297},
            "A3": {"width": 297, "height": 420}} 

def generate_verticle_lab_report(
    worksheet,
    payload,
    coc_id,
    is_cocid: bool = True,
    workbook: Workbook = None,
    region_code=None,
):
    """
    Definition:
        - Generate Verticle Lab Report

    Args:
        - worksheet (Worksheet): Worksheet of excel file output report
        - payload (json): Input paramaters 
        - coc_id (string): Value of cocid (with option COC) or fileid (with option NonCOC)
        - is_cocid (boolean): Paramaters identify as coc or noncoc
        - workbook (Workbook):  Workbook of excel file output report

    Returns:
        - Dataframe: Dataframe after added new column
    """
     
    __lab_name = get_lab_name_report_by_cocid(coc_id=coc_id, is_cocid=is_cocid)
    print(f"__lab_name: {__lab_name}")

    if __lab_name == "" or __lab_name == "R J Hill Laboratories Ltd":
        __lab_name = "Hill"

    try:
        # column A generate
        __columns_templates = get_rows_template_excel(
            payload=payload, coc_id=coc_id, is_cocid=is_cocid, __lab_name=__lab_name
        )

        __row_indexs = 8 + len(__columns_templates) 
        categories = []
        non_categories = []
        counter_non_categories = 0 
        testcode_not_in_db = []
        __columns_templates_remove_non = []

        for item in __columns_templates:
            if item["btestname"] is None and ("w/w" in item["labtestname"]):
                continue
            if item["btestname"] is None:
                item["flg_not_in"] = True
                if not item["categorydescription"] or item["categorydescription"] == "":
                    item["categorydescription"] = ""
                
                testcode_not_in_db.append(item)
                non_category = item["categorydescription"]
                if non_category not in non_categories: 
                    non_categories.append(non_category)
                    if non_category != '':
                        counter_non_categories = counter_non_categories + 1
            else:
                item["flg_not_in"] = False
                __columns_templates_remove_non.append(item)
                
                category = item["categorydescription"]
                if category not in categories:
                    categories.append(category) 
         
        __columns_templates = __columns_templates_remove_non 
        __row_indexs = __row_indexs + len(categories) + counter_non_categories
        
        # generate data for cell column A
        indicated_asbestos_flg, tracking_rows, tracking_rows_non = generate_column_a(
            worksheet=worksheet,
            categories=categories,
            non_categories = non_categories,
            __columns_templates=__columns_templates,
            testcode_not_in_db=testcode_not_in_db,
            is_cocid=is_cocid
        )
        
        if indicated_asbestos_flg:
            __row_indexs += 1
            
        # generate footer column
        __style_notes, start_note, end_note = generate_notes_footer_vertical(
            worksheet=worksheet, payload=payload, row_index=__row_indexs + 1 
        )
         
        __style_notes_logic = {}
        __style_notes_ic = {}
        for style in __style_notes:
            __style_notes_logic[style["businesscriteriacode"]] = style["style"]
            __style_notes_ic[style["businesscriteriacode"]] = style["notecode"]
        __style_mapping_logic = {}
        used_styles = set()
        # get criteria header
        criterias = get_columns_criteria(
            payload=payload, __lab_name=__lab_name, region_code=region_code
        )

        # get criteria datas
        __criteria_datas = get_datas_criteria(
            payload=payload, __lab_name=__lab_name)
        # Generate header criteria xlsx file
        print(f"{__style_notes_ic}")
        __general_criteria_columns = 0
        if criterias and len(criterias) > 0:
            __general_criteria_columns = len(criterias)
            for i in range(__general_criteria_columns):
                column = i + 3
                _criteria_header = criterias[i]
                if _criteria_header["businesscriteriacode"] in __style_notes_logic:
                    __style_mapping_logic[column] = __style_notes_logic[
                        _criteria_header["businesscriteriacode"]
                    ]
                worksheet.insert_cols(column + 1)
                _criterianame = _criteria_header["criterianame"]

                with_cols = 19.44
                if _criteria_header["businesscriteriacode"] in __style_notes_ic:
                    used_styles.add(
                        __style_notes_ic[_criteria_header["businesscriteriacode"]]
                    )
                    worksheet.cell(
                        row=3, column=column
                    ).value = f"{_criterianame}{get_notecode_style(__style_notes_ic[_criteria_header['businesscriteriacode']])}"
                else:
                    worksheet.cell(
                        row=3, column=column).value = f"{_criterianame}"
                worksheet.cell(row=3, column=column).style = "highlight"
                worksheet.cell(row=3, column=column).alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )
                worksheet.merge_cells(
                    start_row=3, start_column=column, end_row=6, end_column=column
                )
                used_styles = generate_criteria_cell_data(
                    col=column,
                    worksheet=worksheet,
                    businesscriteriacode=_criteria_header["businesscriteriacode"],
                    datas=__criteria_datas,
                    __row_indexs=__row_indexs,
                    payload=payload,
                    tracking_rows=tracking_rows,
                    tracking_rows_non=tracking_rows_non,
                    used_styles=used_styles,
                )
                
                print(f"pass here {column}--{used_styles}")

                worksheet.column_dimensions[
                    f"{worksheet.cell(row=3, column= column).column_letter}"
                ].width = with_cols
                # i = i + 1
        # get data and sample
        __samples = get_lab_sample_name_by_cocid(coc_id, is_cocid=is_cocid)

        # count cols sample
        cols = len(__samples)

        # add column max value
        __max_cols = __general_criteria_columns + 3
        worksheet.insert_cols(__max_cols)
        worksheet.cell(row=3, column=__max_cols).value = "Maximum"
        worksheet.cell(row=3, column=__max_cols).style = "highlight"
        worksheet.cell(row=3, column=__max_cols).alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        worksheet.merge_cells(
            start_row=3, start_column=__max_cols, end_row=6, end_column=__max_cols
        )
        generate_maximum_cell_data(
            max_col=__max_cols, worksheet=worksheet, max_rows=__row_indexs
        )

        # Generate header xlsx file
        indicated_asbestos_index = get_asbestos_row(
            worksheet=worksheet, row_indexs=__row_indexs
        )
        columns_rgn_cmp = [3, __general_criteria_columns + 3]
        if __samples and len(__samples) > 0:
            border_right_flg = False
            for i in range(cols):
                column = i + __general_criteria_columns + 4
                __sample = __samples[i]
                if "non" == __sample["depth"]:
                    __sample["depth"] = ""
                    if "_" in __sample["sample_name"]:
                        __sample["depth"] = (
                            __sample["sample_name"]
                            .strip()
                            .split(" ")[0]
                            .split("_")[-1]
                            .split("-")[0]
                            .replace("m", "")
                        )
                    __sample["date"] = get_date_from_sample_name(__sample["sample_name"])
                
                # get width column for header        
                with_cols = len(__sample["sample_id"]) 
                worksheet.cell(
                    row=3, column=column).value = __sample["sample_name"]
                worksheet.cell(
                    row=3, column=column).style = "normal_bolder_header"
                if __sample["depth"] != "" and "non" != __sample["depth"]:
                    worksheet.cell(row=4, column=column).value = str(
                        round(float(__sample["depth"]), 2)
                    )
                worksheet.cell(
                    row=4, column=column).style = "normal_bolder_header"
                worksheet.cell(row=5, column=column).value = __sample["strata"]
                worksheet.cell(
                    row=5, column=column).style = "normal_bolder_header"
                worksheet.cell(row=6, column=column).value = __sample["date"]
                worksheet.cell(
                    row=6, column=column).style = "normal_bolder_header"

                # set with size of columns
                worksheet.column_dimensions[
                    f"{worksheet.cell(row=3, column= column).column_letter}"
                ].width = with_cols

                # get mapped data for file
                __datas = get_ttcl_labfiledetail_by_cocid_barcode(
                    coc_id=coc_id, barcode=__sample["barcode"], is_cocid=is_cocid
                )

                # generate data for col
                border_right_flg = False
                if i == cols - 1:
                    border_right_flg = True
                generate_cell_data(
                    col=column,
                    worksheet=worksheet,
                    sample_id=__sample["sample_id"],
                    samples=__datas,
                    row_indexs=__row_indexs,
                    columns_rgn_cmp=columns_rgn_cmp,
                    style_mapping_logic=__style_mapping_logic,
                    tracking_rows=tracking_rows,
                    tracking_rows_non=tracking_rows_non,
                    workbook=workbook,
                    border_right_flg=border_right_flg,
                )
                
                if indicated_asbestos_index != -1:
                    worksheet.row_dimensions[indicated_asbestos_index].height = 30
                    generate_cell_data_asbestos(
                        col=column,
                        row=indicated_asbestos_index,
                        worksheet=worksheet,
                        workbook=workbook,
                    )
                i = i + 1

            # caculator value maximum cell
            generate_value_maximum_cell_data(
                max_col=__max_cols,
                worksheet=worksheet,
                start_col=__max_cols + 1,
                end_col=__max_cols + cols,
                max_rows=__row_indexs,
                tracking_rows=tracking_rows,
                tracking_rows_non=tracking_rows_non,
            )
        used_styles = list(used_styles)
        for k in range(end_note - 1, start_note - 1, -1):
            note_text = worksheet.cell(row=k, column=1).value
            if note_text is not None:
                if note_text.strip().split(" ")[0] not in used_styles:
                    worksheet.delete_rows(k, 1)
        pagesize, hori_flg = get_print_area(worksheet)
        
        worksheet.page_setup.paperHeight = f'{PAGESIZE[pagesize]["height"]}mm'
        worksheet.page_setup.paperWidth = f'{PAGESIZE[pagesize]["width"]}mm'
        if hori_flg:
            worksheet.print_options.horizontalCentered = True
        else:
            worksheet.print_options.verticalCentered = True
            
        worksheet.delete_rows(7, 1)
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = True 

        return worksheet
    except Exception as e:
        print(f"generate_xlsx: {e}")
        return None


def generate_column_a(worksheet, categories, non_categories, __columns_templates, testcode_not_in_db, is_cocid : bool = True):
    """
    Definition:
        - Generate column A for category and test code name
    Args:
        - worksheet (Worksheet): Worksheet of excel file output report
        - categories: list test code has category
        - non_categories: list test code has not category
        - __columns_templates: List columns template
        - testcode_not_in_db: list test code not exitsts in database
        - is_cocid (boolean): Paramaters identify as coc or noncoc 

    Returns:
        - indicated_asbestos_flg: (bool) determines whether the report contains asbestos.
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize 
    """
    try:
        ### update checkpoint###
        tracking_rows = {}
        tracking_rows_non = []
        ### update checkpoint###
        i = 1
        j = 2
        data = []
        ### update checkpoint###
        data_key = []
        ### update checkpoint###
        data_non = []
        data_key_non = []
        indicated_asbestos_flg = False
        found_flg = False
        maxM = -1
        
       
        for item in categories:
            data = []
            data_key = []
            worksheet.cell(row=5 + i + j, column=1).value = str(item)
            worksheet.cell(
                row=5 + i + j, column=1).style = "define_style_cell_a_bolder"
            print(f"before {__columns_templates}")
            for column in __columns_templates:
                if column["categorydescription"] == item:
                    data.append(column["labtestname"])
                    ### update checkpoint###
                    data_key.append(column["testcode_key"]) 
                    
            print(f"start {item} {data} {data_key} {len(data) == len(data_key)}")

            # ###update checkpoint###
            if (
                not found_flg
                and "Asbestos Fibres/Fine (w/w %)" in data
                and "Asbestos as ACM (w/w%)" in data
            ):
                mAF = data.index("Asbestos Fibres/Fine (w/w %)")
                mAA = data.index("Asbestos as ACM (w/w%)")
                maxM = max(mAF, mAA) + 1
                data.insert(maxM, "Indicated Asbestos Works Control⁴")
                data_key.insert(maxM, "Indicated Asbestos Works Control⁴_")
                indicated_asbestos_flg = True
                if maxM != -1 and mAF > -1 and mAA > -1:
                    found_flg = True
            # write test value
            i = i + 1
            ### update checkpoint###
            for k in range(0, len(data)):
                worksheet.cell(row=5 + i + j, column=1).value = str(data[k])
                if data[k] == "Indicated Asbestos Works Control⁴":
                    worksheet.cell(
                        row=5 + i + j, column=1
                    ).style = "normal_left"
                else:
                    worksheet.cell(row=5 + i + j, column=1).style = "normal_left"
                tracking_rows[5 + i + j] = data_key[k]
                j = j + 1
            ### update checkpoint###

            print(f"end {item}")
        
        print(f'non_categories: {non_categories}')
        i = i + 1
        for item in non_categories:  
            if item != '':
                worksheet.cell(row=5 + i + j, column=1).value = str(item)
                worksheet.cell(
                    row=5 + i + j, column=1).style = "define_highlight_bold_style_left" 
                i = i + 1
                
            data_non = []
            data_key_non = []
            # get testcode_not in db by category
            for column in testcode_not_in_db:
                if column["categorydescription"] == item:
                    data_non.append(column["labtestname"])
                    ### update checkpoint###
                    data_key_non.append(column["testcode_key"])
 
            # write column A for non_testcode in db
            for k in range(0, len(data_non)):  
                tracking_rows_non.append(5 + i + j)
                worksheet.cell(row=5 + i + j, column=1).value = str(data_non[k])
                worksheet.cell(
                    row=5 + i + j, column=1
                ).style = "define_highlight_non_style_left"
                tracking_rows[5 + i + j] = data_key_non[k]
                j = j + 1 
                
            print(f'tracking_rows: {tracking_rows}')
        return indicated_asbestos_flg, tracking_rows, tracking_rows_non
    except Exception as e:
        print(f"Error generate_column_a: {e}")


def generate_criteria_cell_data(
    col,
    worksheet,
    businesscriteriacode,
    datas: list,
    __row_indexs,
    payload,
    tracking_rows,
    tracking_rows_non,
    used_styles,
):
    """
    Definition:
        - Function to generate criteria cell data 
        
    Args:
        - col: Colum index of criteria
        - worksheet (Worksheet): Worksheet of excel file output report
        - businesscriteriacode: bussiness criteria code value
        - datas: List criteria data of bussiness criteria code
        - __row_indexs: Maximum row index of criteria
        - payload: Paramater input value
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize 
        - used_styles: List style hava used
        
    Returns:
        - indicated_asbestos_flg: (bool) determines whether the report contains asbestos.
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize 
    """
    used_styles_copy = used_styles.copy()
    try:
        col_criteria_datas = []
        for item in datas:
            if item["businesscriteriacode"].upper() == businesscriteriacode.upper():
                col_criteria_datas.append(item)

        cells = worksheet["A7":f"A{__row_indexs}"]
        for item in col_criteria_datas:
            index_indicated_asbestos = -1
            for c1 in cells:
                if c1[0].value == "Indicated Asbestos Works Control⁴":
                    index_indicated_asbestos = c1[0].row
                if c1[0].row in tracking_rows:
                    worksheet.cell(row=c1[0].row, column=col).style = "normal"
                    if (
                        businesscriteriacode + "_" + tracking_rows[c1[0].row]
                        == item["bussiness_testcode_key"]
                    ):
                        result = item["testvalue"]
                        if result and result != "NULL":
                            if "Logic" in result:
                                result = apply_smart_criteria(
                                    result,
                                    item["businesscriteriacode"],
                                    item["testcode"],
                                    payload,
                                )
                                if item["notecode"]:
                                    used_styles.add(item["notecode"])
                                    result = result + get_notecode_style(
                                        item["notecode"]
                                    )
                                worksheet.cell(
                                    row=c1[0].row, column=col).value = result
                            else: 
                                # change prefix value criteria
                                __prefix_value = ''
                                if '<' in item["testvalue"]:
                                        __prefix_value = '<'
                                        item["testvalue"] = item["testvalue"].replace('<','')
                                        
                                if '>' in item["testvalue"]:
                                    __prefix_value = '>'
                                    item["testvalue"] = item["testvalue"].replace('>','') 
                                    
                                if item["notecode"]:
                                    used_styles.add(item["notecode"])   
                                    # check cell value of test value. If is nummeric need to format comma for value
                                    if item["testvalue"].isnumeric():  
                                        __value ="{:,}".format(float(item["testvalue"])) 
                                        item["testvalue"] = str(__value).replace('.0','') + get_notecode_style(item["notecode"])
                                    else: 
                                        item["testvalue"] = item["testvalue"] + get_notecode_style(item["notecode"])
                                else:
                                    if item["testvalue"].isnumeric():  
                                        __value ="{:,}".format(float(item["testvalue"])) 
                                        item["testvalue"] = str(__value).replace('.0','')
                                     
                                # write value for cell of criteria data 
                                item["testvalue"] = __prefix_value + item["testvalue"]
                                worksheet.cell(row=c1[0].row, column=col).value = item["testvalue"]
                                
                # If row data is not in tracking row set format hight light style for them.                
                if c1[0].row in tracking_rows_non:
                    worksheet.cell(
                        row=c1[0].row, column=col
                    ).style = "define_highlight_non_style"
        if index_indicated_asbestos != -1:
            worksheet.cell(row=index_indicated_asbestos,
                           column=col).value = "-"
            worksheet.cell(row=index_indicated_asbestos,
                           column=col).style = "normal"
        return used_styles
    except Exception as e:
        print(f"on error {e}")
        return used_styles_copy


def generate_maximum_cell_data(max_col, worksheet, max_rows):
    """
    Definition:
        - Function to generate maximum column data 
        
    Args:
        - max_col: Colum index of maximum
        - worksheet (Worksheet): Worksheet of excel file output report
        - max_rows: Max row of data
        
    Returns:
        - None
    """
    try:
        for i in range(9, max_rows):
            worksheet.cell(i, max_col).value = ""
            worksheet.cell(i, max_col).style = "normal_no_color_no_border"
    except Exception as e:
        print(f"Error generate_maximum_cell_data: {e}")


def generate_value_maximum_cell_data(
    max_col, worksheet, start_col, end_col, max_rows, tracking_rows, tracking_rows_non
):
    """
    Definition:
        - Function to generate criteria cell data 
        
    Args:
        - max_col: Colum index of maximum
        - worksheet (Worksheet): Worksheet of excel file output report
        - start_col: bussiness criteria code value
        - end_col: List criteria data of bussiness criteria code
        - max_rows: Maximum row index of criteria 
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize  
        
    Returns:
        - None
    """
    try:
        for i in range(8, max_rows + 1):
            worksheet.cell(i, max_col).style = "max_cell_style"
            if worksheet.cell(i, 1).value == "Indicated Asbestos Works Control⁴":
                continue
            if i in tracking_rows_non:
                worksheet.cell(i, max_col).style = "define_highlight_non_style"
            if i not in tracking_rows:
                worksheet.cell(i, max_col).value = ""
                worksheet.cell(i, max_col).style = "normal_no_color_no_border"
                continue
            max_value = 0
            max_value_text = ''
            cell_max = None 
            lor = 0
            empty_value = 0
            underlined = 0
            max_j = -1
            for j in range(start_col, end_col + 1):
                _value = worksheet.cell(i, j).value
                if _value is None:
                    empty_value = empty_value + 1
                else:
                    _value_numberic = _value.replace(",", "").replace(".", "").replace("●", "").replace("■", "").replace("▲", "") 
                    if _value_numberic.isnumeric() and max_value < float(_value.replace(",", "").replace("●", "").replace("■", "").replace("▲", "")):
                        max_value = float(_value.replace(",", "").replace("●", "").replace("■", "").replace("▲", ""))
                        max_value_text =_value.replace("●", "").replace("■", "").replace("▲", "")
                        cell_max = copy(worksheet.cell(i, j)._style) 
                        max_j = j
                    else:
                        if "<" in _value or "detected" in _value.lower():
                            lor = lor + 1
                        else:
                            underlined = underlined + 1
                            
            if max_value > 0:
                worksheet.cell(i, max_col).value = max_value_text
                worksheet.cell(i, max_col)._style = cell_max 
                
                # set style for max column in tracking row
                if i  in tracking_rows and i not in tracking_rows_non:
                    worksheet.cell(i, max_col).fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
                    
                if (
                    max_j != -1
                    and worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(i + 2, max_col).value = worksheet.cell(
                        i + 2, max_j
                    ).value
            elif lor > 0:
                worksheet.cell(i, max_col).value = "<LoR"
                if (
                    worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(i + 2, max_col).value = "-"
            elif underlined > 0:
                worksheet.cell(i, max_col).value = "-"
                if (
                    worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(i + 2, max_col).value = "-"
            else:
                worksheet.cell(i, max_col).style = "max_cell_style"
            # else:
            #     worksheet.cell(i, max_col).value = '-'
    except Exception as e:
        print(f"Error generate_value_maximum_cell_data: {e}")


def generate_cell_data(
    col,
    worksheet,
    sample_id,
    samples: list,
    row_indexs,
    columns_rgn_cmp,
    style_mapping_logic,
    tracking_rows,
    tracking_rows_non,
    workbook,
    border_right_flg=False,
):
    """
    Definition:
        - Function to generate cell data for sample
        
    Args:
        - col: Colum index of maximum
        - worksheet (Worksheet): Worksheet of excel file output report
        - sample_id: Sample Id value
        - samples: List sample data
        - row_indexs: Row index of sample
        - columns_rgn_cmp:Value of criteria
        - style_mapping_logic: List style note of cell criteria 
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize  
        - workbook: (workbook) Excel file of result
        - border_right_flg: (bool) Border right
        
    Returns:
        - None
    """
    try:
        col_samples = [] 
        for item in samples:
            if item["samplename"].upper() == sample_id.upper():
                col_samples.append(item)
                 
        cells = worksheet["A7":f"A{row_indexs + 1}"]
        
        if len(col_samples) > 0:
            for item in col_samples:
                for c1 in cells:
                    if c1[0].row in tracking_rows:
                        style_custom = "normal_no_color"
                        symbols = ""
                        if (
                            c1[0].value
                            in [
                                "Asbestos Fibres/Fine (w/w %)",
                                "Asbestos as ACM (w/w%)",
                            ]
                            and c1[0].value in item["testcode"]
                        ):
                            if item["result"] is not None:
                                worksheet.cell(
                                    row=c1[0].row, column=col).value = item["result"]
                                worksheet.cell(
                                    row=c1[0].row, column=col
                                ).style = style_custom
                                if item["resultasnumeric"] is not None:
                                    rs = float(item["resultasnumeric"])
                                    for k in range(
                                        columns_rgn_cmp[0], columns_rgn_cmp[1]
                                    ):
                                        try:
                                            criteria_val = worksheet.cell(
                                                row=c1[0].row, column=k
                                            ).value
                                            if (
                                                "<" in item["result"]
                                                or criteria_val is None
                                                or "-" == criteria_val
                                                or "ND" == criteria_val
                                                or criteria_val == "LoR"
                                            ):
                                                continue
                                            criteria_val = clean_value(
                                                criteria_val)
                                            criteria_val = float(criteria_val)
                                            if rs >= criteria_val:
                                                try:
                                                    if (
                                                        "Black"
                                                        in style_mapping_logic[k]
                                                    ):
                                                        print(f'get_notecode_style(style_mapping_logic[k]): {get_notecode_style(style_mapping_logic[k])}')
                                                        worksheet.cell(
                                                            row=c1[0].row, column=col
                                                        ).value = item[
                                                            "result"
                                                        ] + get_notecode_style(
                                                            style_mapping_logic[k]
                                                        )
                                                    else:
                                                        style_custom += f'define_{"_".join(style_mapping_logic[k].lower().split(" "))}_style' 
                                                except:
                                                    print(
                                                        f"here {style_mapping_logic[k]}"
                                                    )
                                        except Exception as e:
                                            pass
                                    if style_custom != "":
                                        #print(f'style_custom: {style_custom}')
                                        worksheet.cell(
                                            row=c1[0].row, column=col
                                        ).style = custom_style(style_custom, workbook)
                                break
                        
                        # mapping data for cell excel
                        if (item["testcode"] is not None and (item["testcode"] + "_" + item["labcategoryname"]  == tracking_rows[c1[0].row])):
                            if item["result"] is not None:
                                worksheet.cell(row=c1[0].row, column=col).value = item[
                                    "result"
                                ]
                                worksheet.cell(
                                    row=c1[0].row, column=col
                                ).style = style_custom
                                worksheet.cell(
                                    row=c1[0].row, column=col
                                ).number_format = "0.00"
                                # print(f'tracking {item["resultasnumeric"]}')
                                
                                if item["resultasnumeric"] is not None:
                                    rs = float(item["resultasnumeric"])
                                    for k in range(
                                        columns_rgn_cmp[0], columns_rgn_cmp[1]
                                    ):
                                        try:
                                            criteria_val = worksheet.cell(
                                                row=c1[0].row, column=k
                                            ).value
                                            if (
                                                "<" in item["result"]
                                                or criteria_val is None
                                                or "-" == criteria_val
                                                or "ND" == criteria_val
                                                or criteria_val == "LoR"
                                            ):
                                                continue
                                            criteria_val = clean_value(
                                                criteria_val)
                                            criteria_val = float(criteria_val)
                                            if rs >= criteria_val:
                                                try:
                                                    if (
                                                        "Black"
                                                        in style_mapping_logic[k]
                                                    ):
                                                        worksheet.cell(
                                                            row=c1[0].row, column=col
                                                        ).value = item[
                                                            "result"
                                                        ]   
                                                        
                                                        # get symbols for note code style
                                                        symbols+= get_notecode_style(
                                                            style_mapping_logic[k]
                                                        )
                                                    else:
                                                        style_custom += f'define_{"_".join(style_mapping_logic[k].lower().split(" "))}_style' 
                                                except:
                                                    print(
                                                        f"here {style_mapping_logic[k]}"
                                                    )
                                        except Exception as e:
                                            pass
                                    if style_custom != "":
                                        #print(f'style_custom: {style_custom}')
                                        worksheet.cell(
                                            row=c1[0].row, column=col
                                        ).style = custom_style(style_custom, workbook)
                                        
                                    if symbols != "":
                                        print(f'symbols: {symbols}')
                                        worksheet.cell(
                                                            row=c1[0].row, column=col
                                                        ).value += symbols
                                        
                                if c1[0].row in tracking_rows_non:
                                    worksheet.cell(
                                        row=c1[0].row, column=col
                                    ).style = "define_highlight_non_style_center"
                                break 
                    else:
                        if border_right_flg and c1[0].row not in tracking_rows_non:
                            worksheet.cell(
                                row=c1[0].row, column=col
                            ).style = "normal_no_color_border_right"

    except Exception as e:
        print(f"generate_cell_data: {e}")


def generate_cell_data_asbestos(col, row, worksheet, workbook: Workbook = None):
    """
    Definition:
        - Function to generate cell data for asbestos
        
    Args:
        - col: Colum index of cell asbestos
        - row: Row index of cell asbestos
        - worksheet (Worksheet): Worksheet of excel file output report 
        - workbook: (workbook) Excel file of result 
        
    Returns:
        - None
    """
    try:
        asbestos_value = get_indicated_asbestos_value(
            worksheet.cell(row=row - 2, column=col).value,
            worksheet.cell(row=row - 1, column=col).value,
        )
        worksheet.cell(row=row, column=col).value = asbestos_value
        worksheet.cell(
            row=row, column=col).style = "normal_no_color"
        if (
            asbestos_value == "-"
            and worksheet.cell(row=row - 1, column=col).value is None
            and worksheet.cell(row=row - 2, column=col).value is None
        ):
            worksheet.cell(row=row - 1, column=col).value = "-"
            worksheet.cell(row=row - 2, column=col).value = "-"
            worksheet.cell(row=row - 1, column=col).style = "normal_no_color"
            worksheet.cell(row=row - 2, column=col).style = "normal_no_color"
    except Exception as e:
        print(f"generate_cell_data_asbestos: {e}")


def get_asbestos_row(worksheet, row_indexs):
    """
    Definition:
        - Function to get asbestos row
        
    Args: 
        - worksheet (Worksheet): Worksheet of excel file output report 
        - row_indexs: (workbook) Excel file of result 
        
    Returns:
        - row: (int) Row index of asbestos data
    """
    try:
        cells = worksheet["A7":f"A{row_indexs-1}"]
        for c1 in cells:
            if c1[0].value == "Indicated Asbestos Works Control⁴":
                return c1[0].row
        return -1
    except Exception as e:
        print(f"get_asbestos_row: {e}")
        return -1


def get_print_area(worksheet):
    """
    Definition:
        - Function to get print area of xlsx file
        
    Args: 
        - worksheet (Worksheet): Worksheet of excel file output report  
        
    Returns:
        - pg_size: (double) Print page size
        - hori_flg: (bool)  Hrizontal format or Vertical format
    """
    try:
        A4_SIZE = PAGESIZE["A4"]
        A3_SIZE = PAGESIZE["A3"]
        total_width = 0
        for letter in range(1, worksheet.max_column + 1):
            total_width += worksheet.column_dimensions[get_column_letter(
                letter)].width
        print(f"for checking width {total_width}")
        hori_flg = False
        pg_size = "A4"
        threshold = 1.2
        if A4_SIZE["width"] < total_width < A4_SIZE["width"] * threshold:
            return pg_size, hori_flg
        if A4_SIZE["width"] < total_width < A4_SIZE["height"] * threshold:
            hori_flg = True
            return pg_size, hori_flg
        if A3_SIZE["width"] < total_width < A3_SIZE["width"] * threshold:
            pg_size = "A3"
            return pg_size, hori_flg
        if A3_SIZE["width"] < total_width < A3_SIZE["height"] * threshold:
            hori_flg = True
            pg_size = "A3"
            return pg_size, hori_flg
    except Exception as e:
        print(f"get_print_area: {e}") 
