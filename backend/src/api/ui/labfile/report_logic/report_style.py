from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime

def auto_format_cell_width(worksheet, start_column):  
    """
    Definition:
        - Function to make all columns best fit

    Inputs:
        - worksheet (worksheet): Active sheet
        - start_column: Start column need to fit with

    Returns:
        None 
    """
    for letter in range(start_column,worksheet.max_column):
        maximum_value = 0
        for cell in worksheet[get_column_letter(letter)]: 
            if cell.value is not None: 
                val_to_check = len(str(cell.value))
                if val_to_check > maximum_value:
                    maximum_value = val_to_check 
            
        worksheet.column_dimensions[get_column_letter(letter)].width = maximum_value

def set_xlsx_styles(workbook: Workbook): 
    """
    Definition:
        - Function to add range new styles into workbook

    Inputs:
        workbook (Workbook): Save file xlxs output

    Returns:
        None 
    """
    # define style for header data
    workbook.add_named_style(define_style_header())
    
    # define style for bolder cell data
    workbook.add_named_style(define_style_cell_bolder()) 
    workbook.add_named_style(define_style_cell_bolder_header()) 
    
    # define style for cell data
    workbook.add_named_style(define_style_cell())
    workbook.add_named_style(define_style_cell_no_color())
    workbook.add_named_style(define_style_cell_no_color_no_border())
    workbook.add_named_style(define_style_cell_no_color_border_right())
    # define style for max cell data
    workbook.add_named_style(define_style_max_cell())   
    
    # define style for cell data
    workbook.add_named_style(define_style_cell_a_bolder())
    
    # define style for cell data
    workbook.add_named_style(define_style_cell_a_value())
    
    # define style for cell data Criteria
    workbook.add_named_style(define_style_cell_a_value_criteria())   
    
    # Horizontal style title
    workbook.add_named_style(define_horizontal_samples_style_title_cell())
    
    workbook.add_named_style(define_horizontal_samples_style_value_cell())
    
    workbook.add_named_style(define_red_style())
    workbook.add_named_style(define_bold_style())
    workbook.add_named_style(define_underlined_style())
    workbook.add_named_style(define_dashed_outlined_style())
    workbook.add_named_style(define_grey_shaded_style())
    workbook.add_named_style(define_green_outlined_style())
    workbook.add_named_style(define_italicised_style())
    workbook.add_named_style(define_red_style_left())
    workbook.add_named_style(define_bold_style_left())
    workbook.add_named_style(define_underlined_style_left())
    workbook.add_named_style(define_dashed_outlined_style_left())
    workbook.add_named_style(define_grey_shaded_style_left())
    workbook.add_named_style(define_green_outlined_style_left())
    workbook.add_named_style(define_italicised_style_left())
    workbook.add_named_style(define_highlight_non_style())
    workbook.add_named_style(define_highlight_non_style_left())
    workbook.add_named_style(define_highlight_bold_style_left())

    
    workbook.add_named_style(define_highlight_non_style_center())
    workbook.add_named_style(define_style_left())
    workbook.add_named_style(define_highlight_horizontal_samples_style_value_cell())
    workbook.add_named_style(define_style_cell_asbestos())
    workbook.add_named_style(define_style_cell_asbestos_left())
    workbook.add_named_style(define_style_cell_left())
    workbook.add_named_style(define_footer_note_style_title_cell()) 
    workbook.add_named_style(define_style_unofficial())
    
    workbook.add_named_style(define_black_square_style_left())
    workbook.add_named_style(define_black_circle_style_left())
    workbook.add_named_style(define_black_triangle_style_left())
    workbook.add_named_style(define_style_horizontal_title_no_bold())
    
def define_style_header():
    """
    Definition:
        - Function to create new style highlight

    Inputs:
        input_str (str): Sample name

    Returns:
        highlight: (NamedStyle) return style with name highlight
    """
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(name = 'Calibri' ,bold=True, size=9)
    highlight.alignment = Alignment(horizontal='center',
                vertical='center')
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return highlight


def define_style_horizontal_title_no_bold():
    """
    Definition:
        - Function to create new style define_style_horizontal_title_no_bold

    Inputs:
        input_str (str): Sample name

    Returns:
        highlight: (NamedStyle) return style with name define_style_horizontal_title_no_bold
    """
    highlight = NamedStyle(name="define_style_horizontal_title_no_bold")
    highlight.font = Font(name = 'Calibri' , size=9)
    highlight.alignment = Alignment(horizontal='center',
                vertical='center')
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return highlight


def define_style_cell():
    """
    Definition:
        - Function to create new style normal

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name normal
    """
    normal = NamedStyle(name="normal")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center') 
    bd = Side(style='thin', color="000000")
    normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return normal

def define_style_cell_no_color():
    """
    Definition:
        - Function to create new style normal_no_color

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name normal_no_color
    """
    normal = NamedStyle(name="normal_no_color")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center') 
    bd = Side(style='thin', color="000000")
    normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return normal

def define_style_cell_no_color_border_right():
    """
    Definition:
        - Function to create new style normal_no_color_border_right

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name normal_no_color_border_right
    """
    normal = NamedStyle(name="normal_no_color_border_right")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center') 
    bd = Side(style='thin', color="000000")
    normal.border = Border(right=bd)
    return normal


def define_style_cell_no_color_no_border():
    """
    Definition:
        - Function to create new style normal_no_color_no_border

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name normal_no_color_no_border
    """
    normal = NamedStyle(name="normal_no_color_no_border")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center') 
    return normal

def define_style_cell_left():
    """
    Definition:
        - Function to create new style normal_left

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name normal_left
    """
    normal = NamedStyle(name="normal_left")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='left',
                vertical='center') 
    bd = Side(style='thin', color="000000")
    normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return normal

def define_style_cell_asbestos():
    """
    Definition:
        - Function to create new style define_style_cell_asbestos

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name define_style_cell_asbestos
    """
    normal = NamedStyle(name="define_style_cell_asbestos")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center', wrap_text=True) 
    return normal

def define_style_cell_asbestos_left():
    """
    Definition:
        - Function to create new style define_style_cell_asbestos_left

    Inputs:
        input_str (str): Sample name

    Returns:
        normal: (NamedStyle) return style with name define_style_cell_asbestos_left
    """
    normal = NamedStyle(name="define_style_cell_asbestos_left")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='left',
                vertical='center', wrap_text=True)
    normal.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return normal

def define_style_cell_bolder():
    """
    Definition:
        - Function to create new style normal_bolder

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name normal_bolder
    """
    normal_bolder = NamedStyle(name="normal_bolder")
    normal_bolder.font = Font(name = 'Calibri', size=9)
    normal_bolder.alignment = Alignment(horizontal='center',
                vertical='center')
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return normal_bolder

def define_style_cell_bolder_header():
    """
    Definition:
        - Function to create new style normal_bolder_header

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name normal_bolder_header
    """
    normal_bolder = NamedStyle(name="normal_bolder_header")
    normal_bolder.font = Font(name = 'Calibri', size=9)
    normal_bolder.alignment = Alignment(horizontal='center',
                vertical='center')
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal_bolder.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    return normal_bolder

def define_style_max_cell():
    """
    Definition:
        - Function to create new style max_cell_style

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name max_cell_style
    """
    normal = NamedStyle(name="max_cell_style")
    normal.font = Font(name = 'Calibri', size=9)
    normal.alignment = Alignment(horizontal='center',
                vertical='center')  
    normal.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
    bd = Side(style='thin', color="000000")
    normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return normal

def define_style_cell_a_bolder():
    """
    Definition:
        - Function to create new style define_style_cell_a_bolder

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name define_style_cell_a_bolder
    """
    normal_bolder = NamedStyle(name="define_style_cell_a_bolder")
    normal_bolder.font = Font(name = 'Calibri',bold=True, size=9)
    normal_bolder.alignment = Alignment(horizontal='left',
                vertical='center')  
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd)
    return normal_bolder

def define_style_cell_a_value():
    """
    Definition:
        - Function to create new style define_style_cell_a_value

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name define_style_cell_a_value
    """
    normal_bolder = NamedStyle(name="define_style_cell_a_value")
    normal_bolder.font = Font(name = 'Calibri', size=9)  
    
    return normal_bolder

def define_style_cell_a_value_criteria():
    """
    Definition:
        - Function to create new style define_style_cell_a_value_criteria

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name define_style_cell_a_value_criteria
    """
    normal_bolder = NamedStyle(name="define_style_cell_a_value_criteria")
    normal_bolder.font = Font(name = 'Calibri', size=18)  
    
    return normal_bolder

def define_horizontal_samples_style_title_cell():
    """
    Definition:
        - Function to create new style horizontal_title_style

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name horizontal_title_style
    """
    normal_bolder = NamedStyle(name="horizontal_title_style")
    normal_bolder.font = Font(name = 'Calibri', bold=True, size=9)
    normal_bolder.alignment = Alignment(horizontal='left',
                vertical='center')
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal_bolder.fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')


    return normal_bolder

def define_horizontal_samples_style_value_cell():
    """
    Definition:
        - Function to create new style horizontal_samples_style_value

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name horizontal_samples_style_value
    """
    normal_bolder = NamedStyle(name="horizontal_samples_style_value")
    normal_bolder.font = Font(name = 'Calibri', size=9)
    normal_bolder.alignment = Alignment(horizontal='center', vertical='center')
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return normal_bolder

def define_highlight_horizontal_samples_style_value_cell():
    """
    Definition:
        - Function to create new style define_highlight_horizontal_samples_style_value_cell

    Inputs:
        input_str (str): Sample name

    Returns:
        normal_bolder: (NamedStyle) return style with name define_highlight_horizontal_samples_style_value_cell
    """
    
    normal_bolder = NamedStyle(name="define_highlight_horizontal_samples_style_value_cell")
    normal_bolder.font = Font(name = 'Calibri', size=9)
    normal_bolder.alignment = Alignment(horizontal='center', vertical='center')
    bd = Side(style='thin', color="000000")
    normal_bolder.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    normal_bolder.fill = PatternFill(fill_type='solid',start_color='FFA500',end_color='FFA500')
    return normal_bolder

def define_red_style():
    """
    Definition:
        - Function to create new style define_bold_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_bold_style
    """
    ns = NamedStyle(name="define_red_style")
    ns.font = Font(name = 'Calibri', size=9, color='FF0000')
    ns.alignment = Alignment(horizontal='center', vertical='center')
    return ns    

def define_bold_style():
    """
    Definition:
        - Function to create new style define_bold_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_bold_style
    """
    ns = NamedStyle(name="define_bold_style")
    ns.font = Font(name = 'Calibri', size=9, bold=True)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    return ns 

def define_underlined_style():
    """
    Definition:
        - Function to create new style define_underlined_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_underlined_style
    """
    ns = NamedStyle(name="define_underlined_style")
    ns.font = Font(name = 'Calibri', size=9, underline='single')
    ns.alignment = Alignment(horizontal='center', vertical='center')
    return ns

def define_italicised_style():
    """
    Definition:
        - Function to create new style define_italicised_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_italicised_style
    """
    ns = NamedStyle(name="define_italicised_style")
    ns.font = Font(name = 'Calibri', size=9, italic=True)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    return ns 

def define_dashed_outlined_style():
    """
    Definition:
        - Function to create new style define_dashed_outlined_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_dashed_outlined_style
    """
    ns = NamedStyle(name="define_dashed_outlined_style")
    ns.font = Font(name = 'Calibri', size=9)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    bd = Side(style='mediumDashed', color='000000')
    ns.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return ns

def define_grey_shaded_style():
    """
    Definition:
        - Function to create new style define_grey_shaded_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_grey_shaded_style
    """
    ns = NamedStyle(name="define_grey_shaded_style")
    ns.alignment = Alignment(horizontal='center', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    ns.fill = PatternFill(fill_type='solid',start_color='808080',end_color='808080')
    return ns

def define_red_style_left():
    """
    Definition:
        - Function to create new style define_red_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_red_style_left
    """
    ns = NamedStyle(name="define_red_style_left")
    ns.font = Font(name = 'Calibri', size=9, color='FF0000')
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns    

def define_bold_style_left():
    """
    Definition:
        - Function to create new style define_bold_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_bold_style_left
    """ 
    ns = NamedStyle(name="define_bold_style_left")
    ns.font = Font(name = 'Calibri', size=9, bold=True)
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns 

def define_underlined_style_left():
    """
    Definition:
        - Function to create new style define_underlined_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_underlined_style_left
    """ 
    ns = NamedStyle(name="define_underlined_style_left")
    ns.font = Font(name = 'Calibri', size=9, underline='single')
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns

def define_italicised_style_left():
    """
    Definition:
        - Function to create new style define_italicised_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_italicised_style_left
    """ 
    ns = NamedStyle(name="define_italicised_style_left")
    ns.font = Font(name = 'Calibri', size=9, italic=True)
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns 

def define_dashed_outlined_style_left():
    """
    Definition:
        - Function to create new style define_dashed_outlined_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_dashed_outlined_style_left
    """ 
    ns = NamedStyle(name="define_dashed_outlined_style_left")
    ns.font = Font(name = 'Calibri', size=9)
    ns.alignment = Alignment(horizontal='left', vertical='center')
    bd = Side(style='mediumDashed', color='000000')
    ns.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return ns

def define_grey_shaded_style_left():
    """
    Definition:
        - Function to create new style define_grey_shaded_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_grey_shaded_style_left
    """ 
    ns = NamedStyle(name="define_grey_shaded_style_left")
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    ns.fill = PatternFill(fill_type='solid',start_color='808080',end_color='808080')
    return ns

def define_highlight_non_style():
    """
    Definition:
        - Function to create new style define_highlight_non_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_highlight_non_style
    """ 
    ns = NamedStyle(name="define_highlight_non_style")
    ns.alignment = Alignment(horizontal='center', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    ns.fill = PatternFill(fill_type='solid',start_color='FFA500',end_color='FFA500')
    return ns

def define_highlight_non_style_left():
    """
    Definition:
        - Function to create new style define_highlight_non_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_highlight_non_style_left
    """ 
    ns = NamedStyle(name="define_highlight_non_style_left")
    ns.font = Font(name = 'Calibri', size=9)
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.fill = PatternFill(fill_type='solid',start_color='FFA500',end_color='FFA500') 
    return ns

def define_highlight_bold_style_left():
    """
    Definition:
        - Function to create new style define_highlight_bold_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_highlight_bold_style_left
    """ 
    ns = NamedStyle(name="define_highlight_bold_style_left")
    ns.font = Font(name = 'Calibri', size=9, bold=True)
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.fill = PatternFill(fill_type='solid',start_color='FFA500',end_color='FFA500') 
    return ns


def define_highlight_non_style_center():
    """
    Definition:
        - Function to create new style define_highlight_non_style_center

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_highlight_non_style_center
    """ 
    ns = NamedStyle(name="define_highlight_non_style_center")
    ns.font = Font(name = 'Calibri', size=9)
    ns.alignment = Alignment(horizontal='center', vertical='center')
    ns.fill = PatternFill(fill_type='solid',start_color='FFA500',end_color='FFA500') 
    return ns

def define_style_left():
    """
    Definition:
        - Function to create new style define_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_style_left
    """ 
    ns = NamedStyle(name="define_style_left")
    ns.font = Font(name = 'Calibri', size=9) 
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns 

def define_green_outlined_style():
    """
    Definition:
        - Function to create new style define_green_outlined_style

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_green_outlined_style
    """ 
    ns = NamedStyle(name="define_green_outlined_style")
    ns.alignment = Alignment(horizontal='center', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    bd = Side(style='medium', color="4FAA2A")
    ns.border = Border(left=bd, top=bd, right=bd, bottom=bd) 
    return ns

def define_green_outlined_style_left():
    """
    Definition:
        - Function to create new style define_green_outlined_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_green_outlined_style_left
    """ 
    ns = NamedStyle(name="define_green_outlined_style_left")
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    bd = Side(style='medium', color="4FAA2A")
    ns.border = Border(left=bd, top=bd, right=bd, bottom=bd) 
    return ns

def custom_style(style, workbook: Workbook, horizontal='center'):
    """
    Definition:
        - Function to create new style  style + datetime.now().strftime("%Y%m%d%H%M%S%f")

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name  style + datetime.now().strftime("%Y%m%d%H%M%S%f")
    """ 
    style = style + datetime.now().strftime("%Y%m%d%H%M%S%f")
    color = "000000" 
    bold = False
    italic = False
    fill = None
    is_underline = False
    
    bd = Side(style='thin', color="000000")
    
    if "define_red_style" in style:
        color = 'FF0000'
        
    if "define_bold_style" in style:
        bold = True
        
    if "define_underlined_style" in style: 
        is_underline = True
        
    if "define_italicised_style" in style:
        italic = True
        
    if "define_dashed_outlined_style" in style:
        bd = Side(style='mediumDashed', color='000000')
        
    if "define_grey_shaded_style" in style:
        fill = PatternFill(fill_type='solid',start_color='808080',end_color='808080')
        
    if "define_green_outlined_style" in style:
        bd = Side(style='medium', color="4FAA2A") 
    
    ns = NamedStyle(name=style)
    ns.alignment = Alignment(horizontal=horizontal, vertical='center')
    
    if bd is not None:
        ns.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        
    if fill is not None:
        ns.fill = fill
        
    if is_underline:
        ns.font = Font(name = 'Calibri', size=9, italic=italic, bold=bold, color=color, underline='single')
    else:
        ns.font = Font(name = 'Calibri', size=9, italic=italic, bold=bold, color=color)
        
    if style not in workbook.named_styles:
        workbook.add_named_style(ns)
    return style

def define_footer_note_style_title_cell():
    """
    Definition:
        - Function to create new style note_style_title_cell

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name note_style_title_cell
    """ 
    normal_bolder = NamedStyle(name="note_style_title_cell")
    normal_bolder.font = Font(name = 'Calibri', bold=True, size=9)
    normal_bolder.alignment = Alignment(horizontal='left',
                vertical='center')  
    
    return normal_bolder

def define_style_unofficial():
    """
    Definition:
        - Function to create new style define_style_unofficial

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_style_unofficial
    """ 
    ns = NamedStyle(name="define_style_unofficial")
    ns.font = Font(name = 'Calibri', bold=True, size=36, color="ff0000") 
    ns.alignment = Alignment(horizontal='left', vertical='center')
    return ns
 
def define_black_square_style_left():
    """
    Definition:
        - Function to create new style define_black_square_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_black_square_style_left
    """ 
    ns = NamedStyle(name="define_black_square_style_left")
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    return ns

def define_black_circle_style_left():
    """
    Definition:
        - Function to create new style define_black_circle_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_black_circle_style_left
    """ 
    ns = NamedStyle(name="define_black_circle_style_left")
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    return ns

def define_black_triangle_style_left(): 
    """
    Definition:
        - Function to create new style define_black_triangle_style_left

    Inputs:
        input_str (str): Sample name

    Returns:
        ns: (NamedStyle) return style with name define_black_triangle_style_left
    """ 
    ns = NamedStyle(name="define_black_triangle_style_left")
    ns.alignment = Alignment(horizontal='left', vertical='center')
    ns.font = Font(name = 'Calibri', size=9)
    return ns

