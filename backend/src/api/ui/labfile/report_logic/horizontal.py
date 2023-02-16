from openpyxl import Workbook
from src.api.ui.labfile.report_logic.report_common import get_date_from_sample_name
from src.api.ui.labfile.report_logic.report_common import get_lab_name_report_by_cocid, get_indicated_asbestos_value, get_ttcl_labfiledetail_by_cocid_barcode
from src.api.ui.labfile.report_logic.report_style import auto_format_cell_width, custom_style
from src.api.ui.labfile.report_logic.report_common import apply_smart_criteria 
from src.api.ui.labfile.report_logic.report_common import generate_notes_footer_horizontal
from src.api.ui.labfile.report_logic.report_common import get_lab_sample_name_by_cocid, get_rows_template_excel, get_columns_criteria, get_datas_criteria, get_notecode_style, clean_value
from openpyxl.styles import Alignment
from copy import copy
from openpyxl.styles import PatternFill

def generate_horizontal_lab_report(worksheet, payload, coc_id, is_cocid: bool = True, workbook: Workbook = None, region_code=None):
    """
    Definition:
        - This function generates horizontal lab report from input parameters.
    
    Args:
        - worksheet: (worksheet) Active worksheet in Excel file
        - payload: (any) Input paramater from A
        - coc_id: Value of coc_id or file_id
        - is_cocid: If value input is coc then True and if value input of Non-coc is False
        - workbook: (workbook) Workbook of Excel file
        - region_code: Region paramater value
        
    Returns:
        - worksheet : (worksheet) worksheet contains data excel file has been generated.
     """
    __lab_name = get_lab_name_report_by_cocid(coc_id=coc_id, is_cocid=is_cocid)
    print(f'__lab_name: {__lab_name}')

    if __lab_name == '' or __lab_name == 'R J Hill Laboratories Ltd':
        __lab_name = 'Hill'

    # column A generate
    __columns_templates = get_rows_template_excel(
        payload=payload, coc_id=coc_id, is_cocid=is_cocid, __lab_name=__lab_name)
    if __columns_templates is not None and len(__columns_templates) > 0:
        categories = []
        flg_non = False
        testcode_not_in_db = []
        __columns_templates_remove_non = []

        for item in __columns_templates:
            if item["btestname"] is None and ("w/w" in item["labtestname"]):
                continue
            if item["btestname"] is None:
                item["flg_not_in"] = True
                testcode_not_in_db.append(item)
            else:
                item["flg_not_in"] = False
                __columns_templates_remove_non.append(item)
            if not item["categorydescription"] or item["categorydescription"] == "":
                item["categorydescription"] = ""
                flg_non = True
                continue
            category = item["categorydescription"]
            if category not in categories:
                categories.append(category)
        if flg_non:
            categories.append("")
        categories.sort()
        __columns_templates = __columns_templates_remove_non

        indicated_asbestos_flg, tracking_cols, tracking_cols_non = generate_horizontal_header_columns(
            worksheet=worksheet, categories=categories, __columns_templates=__columns_templates, testcode_not_in_db=testcode_not_in_db)
        # get and generate criteria header datas
        # get criteria header
        criterias = get_columns_criteria(
            payload=payload, __lab_name=__lab_name, region_code=region_code)

        # get data and sample
        __samples = get_lab_sample_name_by_cocid(coc_id, is_cocid=is_cocid)
        # generate footer excel file

        __row_index = 8 + len(__samples)
        
        __style_notes, start_note, end_note = generate_notes_footer_horizontal(
            worksheet=worksheet,payload=payload, row_index=__row_index,  criteria_length=len(criterias))
        __style_notes_logic = {}
        __style_notes_ic = {}
        for style in __style_notes:
            __style_notes_logic[style["businesscriteriacode"]] = style["style"]
            __style_notes_ic[style["businesscriteriacode"]] = style["notecode"]
        __style_mapping_logic = {}
        
        used_styles = set()  
        used_styles, __style_mapping_logic = generate_horizontal_criteria_rows(worksheet=worksheet, criterias=criterias, payload=payload,  tracking_cols=tracking_cols, indicated_asbestos_flg=indicated_asbestos_flg,
                                                                               tracking_cols_non=tracking_cols_non, __style_notes_ic=__style_notes_ic, __style_notes_logic=__style_notes_logic, used_styles=used_styles, __lab_name=__lab_name, __style_mapping_logic=__style_mapping_logic)
        used_styles = list(used_styles)
        # for k in range(end_note + 3, start_note + 2, -1):
        #     note_text = worksheet.cell(row=k, column=1).value
        #     if note_text is not None:
        #         if note_text.strip().split(" ")[0] not in used_styles:
        #             worksheet.delete_rows(k, 1)
        # generate data for sample column
        __criteria_length = 0
        if criterias is not None:
            __criteria_length = len(criterias)

        columns_rgn_cmp = [5, __criteria_length + 5]
        generate_horizontal_samples_rows(worksheet=worksheet, samples=__samples, coc_id=coc_id, is_cocid=is_cocid, criteria_length=__criteria_length, tracking_cols=tracking_cols,
                                         tracking_cols_non=tracking_cols_non, columns_rgn_cmp=columns_rgn_cmp, __style_mapping_logic=__style_mapping_logic, workbook=workbook)

        # set auto with columns in report
        auto_format_cell_width(worksheet=worksheet, start_column=5)

        print(f'Run success generate_horizontal_lab_report')

        return worksheet


def generate_horizontal_header_columns(worksheet, categories, __columns_templates, testcode_not_in_db):
    """
    Definition:
        - This function generates horizontal lab report from input parameters.
    
    Args:
        - worksheet: (worksheet) Active worksheet in Excel file
        - categories: (dict) This parameter contains the list of testcodes and testcategory
        - __columns_templates: This parameter contains the list of columns to generate
        - testcode_not_in_db: List of testcodes that exist in the input csv file but not in the criteria's initializ
        
    Returns:
        - indicated_asbestos_flg: (bool) determines whether the report contains asbestos.
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: (dict) list of columns that not exist in criteria initialize 
    """
    try:
        tracking_cols = {}
        tracking_cols_non = []
        i = 0
        j = 0
        data = []
        data_key = []
        data_non = []
        data_key_non = []
        max_height = 0
        indicated_asbestos_flg = False
        found_flg = False
        maxM = -1
        for item in categories:
            for column in __columns_templates:
                if column["categorydescription"] == item:
                    data.append(column)
                    data_key.append(column["testcode_key"])
           
            ### update checkpoint###
            if not found_flg:
                mAF = -1
                mAA = -1
                ii = 0
                while (ii <= len(data) - 1) and (mAF == -1 or mAA == -1):
                    if data[ii]["labtestname"] == "Asbestos as ACM (w/w%)":
                        mAA = ii
                    if data[ii]["labtestname"] == "Asbestos Fibres/Fine (w/w %)":
                        mAF = ii
                    ii += 1
                maxM = max(mAF, mAA) + 1
                if maxM > 0:
                    data.insert(
                        maxM, {"labtestname": "Indicated Asbestos Works Control⁴", "unit": ""})
                    data_key.insert(maxM, "Indicated Asbestos Works Control⁴_")
                    indicated_asbestos_flg = True
                    found_flg = True
            start_column = 5 + j
            end_column = 5 + j + len(data) - 1
            # if indicated_asbestos_flg:
            #     print(f"for checking {maxM}")
            #     end_column += 1
            print(f'Run to generate_horizontal_header_columns> merge_cells')

            worksheet.cell(row=3, column=start_column).alignment = Alignment(horizontal='center', vertical='bottom',
                                                                             wrap_text=False, shrink_to_fit=False)

            worksheet.cell(row=3, column=start_column).value = item
            worksheet.cell(row=3, column=start_column).style = "highlight"

            if end_column > start_column:
                worksheet.merge_cells(
                    start_row=3, start_column=start_column, end_row=3, end_column=end_column)
                print(f'Merge cell if end_column > start_column: success')
            # write test value
            print(f'data generate_horizontal_header_columns: {data}')
            for k in range(0, len(data)):
                # Get max height of column
                __height = len(str(data[k]["labtestname"]))
                if __height > max_height:
                    max_height = __height
                worksheet.cell(row=4, column=5 +
                               j).value = str(data[k]["labtestname"])
                worksheet.cell(row=4, column=5 +
                               j).style = "highlight"
                worksheet.cell(row=4, column=5+j).alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=90,
                                                                        wrap_text=False, shrink_to_fit=False)

                worksheet.cell(row=5, column=5+j).value = str(data[k]["unit"])
                worksheet.cell(row=5, column=5 +
                               j).style = "highlight"
                worksheet.cell(row=5, column=5+j).alignment = Alignment(horizontal='center', vertical='center',
                                                                        wrap_text=False, shrink_to_fit=False)
                tracking_cols[5+j] = data_key[k]
                j = j+1

            i = i + 1
            data = []
            data_key = []
        # set max heigh column
        i = i - 1
        for column in testcode_not_in_db:
            data_non.append(column)
            ### update checkpoint###
            data_key_non.append(column["testcode_key"])
            ### update checkpoint###
       
        ### update checkpoint###
        for k in range(0, len(data_non)):
            tracking_cols_non.append(5+j)
            # Get max height of column
            __height = len(str(data_non[k]["labtestname"]))
            if __height > max_height:
                max_height = __height
            worksheet.cell(row=4, column=5 +
                           j).value = str(data_non[k]["labtestname"])
            worksheet.cell(
                row=4, column=5+j).style = "highlight"
            worksheet.cell(row=4, column=5+j).alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=90,
                                                                    wrap_text=False, shrink_to_fit=False)

            worksheet.cell(row=5, column=5+j).value = str(data_non[k]["unit"])
            worksheet.cell(
                row=5, column=5+j).style = "highlight"
            worksheet.cell(row=5, column=5+j).alignment = Alignment(horizontal='center', vertical='center',
                                                                    wrap_text=False, shrink_to_fit=False)
            tracking_cols[5+j] = data_key_non[k]
            j = j+1
        if len(data_non) > 0:
            worksheet.merge_cells(start_row=3, start_column=(
                5+j-len(data_non)), end_row=3, end_column=5+j-1)
        if max_height > 0:
            worksheet.row_dimensions[4].height = max_height * 4
        print(
            f"showing: {indicated_asbestos_flg}, {tracking_cols}, {tracking_cols_non}")
        return indicated_asbestos_flg, tracking_cols, tracking_cols_non
    except Exception as e:
        print(f'Method generate_horizontal_header_columns error: {e}')
        return None, None, None


def generate_horizontal_criteria_rows(worksheet, criterias, payload, tracking_cols, indicated_asbestos_flg, tracking_cols_non, __style_notes_ic, __style_notes_logic, used_styles, __lab_name, __style_mapping_logic):
    """
    Definition:
        - This function writes criteria header for lab report
    
    Args:
        - worksheet: (worksheet) Active worksheet in Excel file
        - criterias: (dict) This parameter contains the list of testcodes and testcategory of criteria 
        - payload: This parameter contains paramater
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - indicated_asbestos_flg: (bool) determines whether the report contains asbestos
        - tracking_cols_non: dict) list of columns that not exist in criteria initialize 
        - __style_notes_ic: List style note of header criteria
        - __style_notes_logic: List style note of cell criteria
        - used_styles: List style hava used
        - __lab_name: Lab name of report (Hill, Analytica, Eurofill..)
        - __style_mapping_logic:
        
    Returns:
        - used_styles: List style hava used
        - __style_mapping_logic: List style note of cell criteria
    """
    try:
        # Generate header criteria xlsx file
        __general_criteria_rows = 0
        row_index = 0
        if criterias and len(criterias) > 0:
            __general_criteria_rows = len(criterias)
            # i = 0
            for i in range(__general_criteria_rows):
                print(f'row_index = i + 6: {i + 6}')
                row_index = i + 6
                worksheet.insert_rows(row_index)
                _criteria_header = criterias[i]
                __style_mapping_logic[row_index] = __style_notes_logic[_criteria_header["businesscriteriacode"]]
                _criterianame = _criteria_header["criterianame"]
                worksheet.cell(
                    row=row_index, column=1).style = 'horizontal_title_style'
                worksheet.cell(
                    row=row_index, column=1).value = f"{_criterianame}{get_notecode_style(__style_notes_ic[_criteria_header['businesscriteriacode']])}"
                if _criteria_header['businesscriteriacode'] in __style_notes_ic:
                    used_styles.add(
                        __style_notes_ic[_criteria_header['businesscriteriacode']])
                worksheet.merge_cells(
                    start_row=row_index, start_column=1, end_row=row_index, end_column=4)

                # get row information
                current_row = list(worksheet.rows)[row_index-1]
                start_cell = current_row[4]
                end_cell = current_row[-1]

                # genere data for criteria cells
                # get criteria datas
                __criteria_datas = get_datas_criteria(
                    payload=payload, __lab_name=__lab_name)
                __businesscriteriacode = _criteria_header["businesscriteriacode"]
                used_styles = generate_horizontal_criteria_data(start_cell=start_cell, end_cell=end_cell, worksheet=worksheet, businesscriteriacode=__businesscriteriacode, datas=__criteria_datas,
                                                                payload=payload, tracking_cols=tracking_cols, tracking_cols_non=tracking_cols_non, __style_notes_ic=__style_notes_ic, __style_notes_logic=__style_notes_logic, used_styles=used_styles)
                 
            # Write maximum row data
            _maximum_row_index = __general_criteria_rows + 6
            worksheet.insert_rows(_maximum_row_index) 
            worksheet.cell(row=_maximum_row_index, column=1).style = 'horizontal_title_style'
            worksheet.cell(row=_maximum_row_index, column=1).value = "Maximum"    
            worksheet.merge_cells(start_row=_maximum_row_index, start_column=1, end_row=_maximum_row_index, end_column=4)
            
        return used_styles, __style_mapping_logic
    except Exception as e:
        print(f'Method generate_horizontal_criteria_rows error 2: {e}') 

def generate_horizontal_criteria_data(start_cell, end_cell, worksheet, businesscriteriacode, datas: list, payload, tracking_cols, tracking_cols_non, __style_notes_ic, __style_notes_logic, used_styles):
    """
    Definition:
        - This function writes cell data of criteria for lab report
    
    Args:
        - start_cell: Cell start of row data
        - end_cell: Cell end of row data
        - worksheet: (worksheet) Active worksheet in Excel file
        - businesscriteriacode: Value of criteria
        - datas: List initialize data for criteria
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: dict) list of columns that not exist in criteria initialize 
        - __style_notes_logic: List style note of cell criteria
        - used_styles: List style hava used
          
    Returns:
        - used_styles: List style hava used 
    """
    try:
        used_styles_copy = used_styles.copy()
        col_criteria_datas = []
        for item in datas:
            if item["businesscriteriacode"].upper() == businesscriteriacode.upper():
                col_criteria_datas.append(item)
        cells = worksheet['E4': f'{end_cell.column_letter}4']
        for c1 in cells[0]:
            if c1.value == "Indicated Asbestos Works Control⁴":
                worksheet[f'{c1.column_letter}{start_cell.row}'].value = "-"
                worksheet[f'{c1.column_letter}{start_cell.row}'].style = 'define_style_horizontal_title_no_bold'
                continue
            if c1.column in tracking_cols:
                worksheet[f'{c1.column_letter}{start_cell.row}'].style = 'define_style_horizontal_title_no_bold'
                for item in col_criteria_datas: 
                    if businesscriteriacode + "_" + tracking_cols[c1.column] == item["bussiness_testcode_key"]:
                        # print(
                        #     f'generate_horizontal_criteria_data - item : {item}')
                        result = item["testvalue"]
                        if result and result != 'NULL':
                            if "Logic" in result:
                                result = apply_smart_criteria(
                                    result, item["businesscriteriacode"], item["testcode"], payload)
                                if item["notecode"]:
                                    used_styles.add(item["notecode"])
                                    result = result + \
                                        get_notecode_style(item["notecode"])
                                worksheet[f'{c1.column_letter}{start_cell.row}'].value = result
                            else:
                                
                                # change prefix value criteria
                                __prefix_value = ''
                                if '<' in item["testvalue"]:
                                        __prefix_value = '<'
                                        item["testvalue"] = item["testvalue"].replace('<','')
                                        
                                if '>' in item["testvalue"]:
                                    __prefix_value = '>'
                                    item["testvalue"] = item["testvalue"].replace('>','') 
                                    
                                if item["notecode"]:
                                    used_styles.add(item["notecode"])   
                                    # check cell value of test value. If is nummeric need to format comma for value
                                    if item["testvalue"].isnumeric():  
                                        __value ="{:,}".format(float(item["testvalue"])) 
                                        item["testvalue"] = str(__value).replace('.0','') + get_notecode_style(item["notecode"])
                                    else: 
                                        item["testvalue"] = item["testvalue"] + get_notecode_style(item["notecode"])
                                else:
                                    if item["testvalue"].isnumeric():  
                                        __value ="{:,}".format(float(item["testvalue"])) 
                                        item["testvalue"] = str(__value).replace('.0','')
                                     
                                # write value for cell of criteria data 
                                item["testvalue"] = __prefix_value + item["testvalue"]
                                worksheet[f'{c1.column_letter}{start_cell.row}'].value = item["testvalue"]
                                
                                # if item["notecode"]:
                                #     used_styles.add(item["notecode"])
                                #     item["testvalue"] = item["testvalue"] + \
                                #         get_notecode_style(item["notecode"])
                                # worksheet[f'{c1.column_letter}{start_cell.row}'].value = item["testvalue"]
                        break
                if c1.column in tracking_cols_non:
                    worksheet.cell(
                        row=start_cell.row, column=c1.column).style = "define_highlight_horizontal_samples_style_value_cell"
        return used_styles
    except Exception as e:
        print(f'Method generate_horizontal_criteria_data error: {e}')
        return used_styles_copy

def generate_horizontal_samples_rows(worksheet, samples, coc_id, is_cocid, criteria_length, tracking_cols, tracking_cols_non, columns_rgn_cmp, __style_mapping_logic, workbook: Workbook = None):
    """
    Definition:
        - This function writes sample header for lab report
    
    Args: 
        - worksheet: (worksheet) Active worksheet in Excel file
        - samples: List samples of input csv file
        - coc_id: Value of coc_id if is_cocid = True and Value of file_id if is_cocid = False
        - is_cocid: True if case COC and False if case None-COC
        - criteria_length: Counter of criteria have choise.
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: dict) list of columns that not exist in criteria initialize  
        - columns_rgn_cmp: Value of criteria 
        - __style_mapping_logic: List style note of cell criteria
        - workbook: (workbook) Excel file of result
          
    Returns:
        - None
    """
    try:
        if samples is not None and len(samples) > 0:
            i = 0
            indicated_asbestos_index = get_asbestos_column(tracking_cols)
            for sample in samples:
                row_index = 8 + i + criteria_length
                worksheet[f'A{row_index}'].value = sample["sample_name"]
                worksheet[f'A{row_index}'].style = "highlight"
                if "non" == sample["depth"]:
                    sample["depth"] = ""
                    
                    if "_" in sample["sample_name"]:
                        sample["depth"] = sample["sample_name"].strip().split(
                            " ")[0].split("_")[-1].split("-")[0].replace("m", "")
                    
                    sample["strata"] = "non"
                    
                    sample["date"] = get_date_from_sample_name(sample["sample_name"])
                        
                if sample["depth"] is not None and sample["depth"] != "":
                    worksheet[f'B{row_index}'].value = str(round(float(sample["depth"]), 2))
                    
                worksheet[f'B{row_index}'].style = "highlight"

                worksheet[f'C{row_index}'].value = sample["strata"]
                worksheet[f'C{row_index}'].style = "highlight"

                worksheet[f'D{row_index}'].value = sample["date"]
                worksheet[f'D{row_index}'].style = "highlight"

                # generate data for cell
                current_row = list(worksheet.rows)[row_index - 1]
                start_cell = current_row[4]
                end_cell = current_row[-1]

                # get mapped data for file
                datas = get_ttcl_labfiledetail_by_cocid_barcode(
                    coc_id=coc_id, barcode=sample['barcode'], is_cocid=is_cocid)

                generate_horizontal_samples_data_rows(start_cell=start_cell, end_cell=end_cell, worksheet=worksheet, sample_name=sample[
                                                      "sample_id"], datas=datas, tracking_cols=tracking_cols, tracking_cols_non=tracking_cols_non, columns_rgn_cmp=columns_rgn_cmp, __style_mapping_logic=__style_mapping_logic, workbook=workbook)
                if indicated_asbestos_index != -1:
                    generate_cell_data_asbestos(
                        col=indicated_asbestos_index, row=row_index, worksheet=worksheet)
                i = i + 1
               
               
            # generate horizontal maximum data rows
            generate_horizontal_maximum_data(start_col_index= 5, #start_cell.row, 
                                             end_col_index= 5 + len(tracking_cols), #end_cell.row,
                                             counter_samples = len(samples),
                                             maximum_row_index = 6 + criteria_length,
                                             tracking_cols= tracking_cols,
                                             tracking_cols_non= tracking_cols_non,
                                             worksheet=worksheet 
                                            )  
    except Exception as e:
        print(f'Method generate_horizontal_samples_rows erroe: {e}')

def generate_horizontal_samples_data_rows(start_cell, end_cell, worksheet, sample_name, datas, tracking_cols, tracking_cols_non, columns_rgn_cmp, __style_mapping_logic, workbook: Workbook = None):
    """
    Definition:
        - This function writes cell data of sample for lab report
    
    Args: 
        - start_cell: Start cell write data
        - end_cell: End cell write data
        - worksheet: (worksheet) Active worksheet in Excel file
        - sample_name: Sample name input
        - datas: Data of sample name
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: dict) list of columns that not exist in criteria initialize  
        - columns_rgn_cmp: Value of criteria 
        - __style_mapping_logic: List style note of cell criteria
        - workbook: (workbook) Excel file of result
          
    Returns:
        - None
    """
    try:
        col_criteria_datas = []
        for item in datas:
            if item["samplename"].upper() == sample_name.upper():
                col_criteria_datas.append(item)
        cells = worksheet[f'{start_cell.column_letter}4': f'{end_cell.column_letter}4']
        for c1 in cells[0]:

            # print(
            #     f'c1.column_letter - start_cell.row: {c1.column_letter}{start_cell.row}')
            worksheet[f'{c1.column_letter}{start_cell.row}'].style = "horizontal_samples_style_value"
            for item in col_criteria_datas:
                style_custom = "normal_bolder"
                if c1.value in ["Asbestos Fibres/Fine (w/w %)", "Asbestos as ACM (w/w%)"] and c1.value in item["testcode"]:
                    result = item["result"]
                    if result and result != 'NULL':
                        worksheet[f'{c1.column_letter}{start_cell.row}'].value = result
                    if item["resultasnumeric"] is not None:
                        rs = float(item["resultasnumeric"])
                        for k in range(columns_rgn_cmp[0], columns_rgn_cmp[1]):
                            try:
                                criteria_val = worksheet.cell(
                                    row=k, column=c1.column).value
                                if "<" in item["result"] or criteria_val is None or "-" == criteria_val or 'ND' == criteria_val or criteria_val == "LoR":
                                    continue
                                criteria_val = clean_value(criteria_val)
                                criteria_val = float(criteria_val)
                                if rs >= criteria_val:
                                    try:
                                        if "Black" in __style_mapping_logic[k]:
                                            worksheet[f'{c1.column_letter}{start_cell.row}'].value = item["result"] + get_notecode_style(
                                                __style_mapping_logic[k])
                                        else:
                                            style_custom += f'define_{"_".join(__style_mapping_logic[k].lower().split(" "))}_style'
                                            # worksheet.cell(row=c1[0].row,column=col).style = f'define_{"_".join(style_mapping_logic[k].lower().split(" "))}_style'
                                    except:
                                        print(
                                            f"here {__style_mapping_logic[k]}")
                            except Exception as e:
                                pass
                    if style_custom != "":
                        worksheet[f'{c1.column_letter}{start_cell.row}'].style = custom_style(
                            style_custom, workbook)
                    break 
               
                if c1.value is not None and item["testcode"] is not None and (item["testcode"] + "_" + item["labcategoryname"] == tracking_cols[c1.column]):
                    result = item["result"]
                     # define style for symbol of data cell
                    symbols = "" 
                    
                    if result and result != 'NULL':
                        worksheet[f'{c1.column_letter}{start_cell.row}'].value = result
                        
                        print(f'value = result: {result}')
                        if item["resultasnumeric"] is not None:
                            rs = float(item["resultasnumeric"])  
                            for k in range(columns_rgn_cmp[0], columns_rgn_cmp[1]):
                                try:
                                    criteria_val = worksheet.cell(
                                        row=k, column=c1.column).value
                                    if "<" in item["result"] or criteria_val is None or "-" == criteria_val or 'ND' == criteria_val or criteria_val == "LoR":
                                        continue
                                    criteria_val = clean_value(criteria_val)
                                    criteria_val = float(criteria_val)
                                    if rs >= criteria_val:
                                        try:
                                            if "Black" in __style_mapping_logic[k]:  
                                                worksheet[f'{c1.column_letter}{start_cell.row}'].value = item["result"]  
                                                        
                                                # get symbols for note code style
                                                symbols += get_notecode_style(
                                                    __style_mapping_logic[k]
                                                )
                                                        
                                            else:
                                                style_custom += f'define_{"_".join(__style_mapping_logic[k].lower().split(" "))}_style' 
                                        except:
                                            print(
                                                f"here {__style_mapping_logic[k]}")
                                except Exception as e:
                                    pass
                    
                    # Apply custom style for cell data
                    if style_custom != "":
                        worksheet[f'{c1.column_letter}{start_cell.row}'].style = custom_style(
                            style_custom, workbook)
                    
                    # Update symbol value for cell excel
                    if symbols != "":
                        print(f'symbols: {symbols}')
                        worksheet[f'{c1.column_letter}{start_cell.row}'].value += symbols

                    if c1.column in tracking_cols_non:
                        worksheet[f'{c1.column_letter}{start_cell.row}'].style = "define_highlight_horizontal_samples_style_value_cell"

                    break

    except Exception as e:
        print(f'Method generate_horizontal_samples_data_rows erroe: {e}')

def generate_horizontal_maximum_data(start_col_index, end_col_index, maximum_row_index, worksheet, tracking_cols, tracking_cols_non, counter_samples):
    """
    Definition:
        - This function writes cell maximum data for lab report
    
    Args:  
        - start_col_index: Start cell of sample data
        - end_col_index: End cell of sample data
        - maximum_row_index: Index row maximum need to write data
        - worksheet: (worksheet) Active worksheet in Excel file 
        - tracking_cols: (dict) list of columns that exist in criteria initialize
        - tracking_cols_non: dict) list of columns that not exist in criteria initialize  
        - counter_samples: Counter of samples
          
    Returns:
        - None
    """
    try:
        
        # Check value earch cell for each test code
        for i in range(start_col_index, end_col_index):
            #print(f'tracking_cols: {tracking_cols}. tracking_cols: {len(tracking_cols)}')
            worksheet.cell(maximum_row_index, i).style = "max_cell_style"
            if worksheet.cell(i, 1).value == "Indicated Asbestos Works Control⁴":
                continue
           
            if i in tracking_cols_non:
                worksheet.cell(maximum_row_index, i).style = "define_highlight_non_style"
            
            if i not in tracking_cols:
                worksheet.cell(maximum_row_index, i).value = ""
                worksheet.cell(maximum_row_index, i).style = "normal_no_color_no_border"
                continue
            max_value = 0
            max_value_text = ''
            cell_max = None 
            lor = 0
            empty_value = 0
            underlined = 0
            max_j = -1
            # Check value for earch column
            for j in range(maximum_row_index + 1, maximum_row_index + 1 + counter_samples):
                _value = worksheet.cell(j, i).value 
                if _value is None:
                    empty_value = empty_value + 1
                else:
                    _value_numberic = _value.replace(",", "").replace(".", "").replace("●", "").replace("■", "").replace("▲", "") 
                    if _value_numberic.isnumeric() and max_value < float(_value.replace(",", "").replace("●", "").replace("■", "").replace("▲", "")):
                        max_value = float(_value.replace(",", "").replace("●", "").replace("■", "").replace("▲", ""))
                        max_value_text =_value.replace("●", "").replace("■", "").replace("▲", "")
                        cell_max = copy(worksheet.cell(j, i)._style) 
                        max_j = j
                    else:
                        if "<" in _value or "detected" in _value.lower():
                            lor = lor + 1
                        else:
                            underlined = underlined + 1
                            
            if max_value > 0:
                worksheet.cell(maximum_row_index, i).value = max_value_text
                worksheet.cell(maximum_row_index, i)._style = cell_max 
                
                # set style for max column in tracking row
                if i in tracking_cols and i not in tracking_cols_non:
                    worksheet.cell(maximum_row_index, i).fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
                else:
                    worksheet.cell(maximum_row_index, i).style = 'define_highlight_horizontal_samples_style_value_cell'
                    
                if (
                    max_j != -1
                    and worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(maximum_row_index, i + 2).value = worksheet.cell(
                        i + 2, max_j
                    ).value
            elif lor > 0:
                worksheet.cell(maximum_row_index, i).value = "<LoR"
                if i in tracking_cols_non:
                    worksheet.cell(maximum_row_index, i).style = 'define_highlight_horizontal_samples_style_value_cell' 
                if (
                    worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(maximum_row_index, i + 2).value = "-"
                    
            elif underlined > 0:
                worksheet.cell(maximum_row_index, i).value = "-"
                if (
                    worksheet.cell(i, 1).value == "Asbestos Fibres/Fine (w/w %)"
                    and worksheet.cell(i + 2, 1).value
                    == "Indicated Asbestos Works Control⁴"
                ):
                    worksheet.cell(i + 2, maximum_row_index).value = "-"
            else:
                if i in tracking_cols and i not in tracking_cols_non:
                    worksheet.cell(maximum_row_index, i).fill = PatternFill(fill_type='solid',start_color='DDEBF7',end_color='DDEBF7')
                else:
                    worksheet.cell(maximum_row_index, i).style = 'define_highlight_non_style'  
            
    except Exception as e:
        print(f'Method generate_horizontal_maximum_data error: {e}')

def generate_cell_data_asbestos(col, row, worksheet):
    """
    Definition:
        - This function writes cell cell data asbestos for lab report
    
    Args:  
        - col: Column write cell data of asbestos 
        - row: Row write cell data of asbestos  
        - worksheet: (worksheet) Active worksheet in Excel file  
          
    Returns:
        - None
    """
    try:
        asbestos_value = get_indicated_asbestos_value(worksheet.cell(
            row=row, column=col-2).value, worksheet.cell(row=row, column=col-1).value)
        worksheet.cell(row=row, column=col).value = asbestos_value
        worksheet.cell(
            row=row, column=col).style = "horizontal_samples_style_value"
        if asbestos_value == "-" and worksheet.cell(row=row, column=col - 1).value is None and worksheet.cell(row=row, column=col - 2).value is None:
            worksheet.cell(row=row, column=col - 1).value = "-"
            worksheet.cell(row=row, column=col - 2).value = "-"
            worksheet.cell(row=row, column=col - 1).style = "horizontal_samples_style_value"
            worksheet.cell(row=row, column=col - 2).style = "horizontal_samples_style_value"
    except Exception as e:
        print(f'generate_cell_data_asbestos: {e}')


def get_asbestos_column(tracking_cols):
    """
    Definition:
        - This function to get index Asbestos has in tracking column
    
    Args:  
        - tracking_cols: List columns have write on Lab report 
          
    Returns:
        - int: if value > 1 has in tracking_cols else value is -1
    """
    try:
        for k in tracking_cols.keys():
            if "Indicated Asbestos Works Control⁴" in tracking_cols[k]:
                return int(k)
        return -1
    except Exception as e:
        print(f'get_asbestos_col: {e}')
        return -1
